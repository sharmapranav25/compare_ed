"""Scaffold a golden test set JSON from an existing v1 xlsx output.

Usage:
    python -m buysheet_v2.tests.scaffold_golden <vendor_key>

Reads ../BUYSHEET_<vendor>.xlsx + ../files/<vendor>/<doc>.reconciled.txt,
samples N SKUs evenly across the catalog, pre-fills the golden template with
the current (post-bug-fix) extracted values, and writes tests/golden/<vendor>.json.

Every entry has `"_verified": false` — the human reviewer flips this to `true`
after manually verifying the values against the source PDF.

The eval harness only counts entries where `_verified == true`.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent.parent
GOLDEN_DIR = Path(__file__).parent / "golden"
HOLDOUT_DIR = Path(__file__).parent / "holdout"

# Maps the v1 COL indices to v2 schema field names
COL_TO_FIELD = {
    2: "sku",
    3: "mg",
    4: "sg",
    5: "ssg",
    6: "description",
    7: "color",
    8: "standard_color",
    16: "intro_date",
    22: "usd_cost",
    23: "usd_retail",
}

# Vendor key -> (xlsx filename, reconciled txt path relative to files/<dir>)
KNOWN_VENDORS = {
    "nike_ho26": ("BUYSHEET_nike_ho26.xlsx", "files/nike_ho26/HO26_GEL_FW.reconciled.txt"),
    "adidas_premium": ("BUYSHEET_adidas_premium.xlsx",
                       "files/adidas_premium/FW26_ORIGINALS_FTW_PREMIUM_RANGE.reconciled.txt"),
    "converse_ho26": ("BUYSHEET_converse_ho26.xlsx",
                      "files/converse_ho26/KITH_CONVERSE_NHBD_HO26_CATALOG.reconciled.txt"),
    "mizuno_ss27": ("BUYSHEET_mizuno_ss27.xlsx",
                    "files/mizuno_ss27/SS27_MIZUNO_SPORTSTYLE_USA_DECK_Tier1.reconciled.txt"),
    "hoka_sp27": ("BUYSHEET_hoka_sp27.xlsx",
                  "files/hoka_sp27/HOKA_SP27_PINNACLE_LOOK_BOOK.reconciled.txt"),
}

# Which vendors are tune-against (golden/) vs holdout/ — the holdouts are
# locked before any Phase 1+ code is written, and are the ship gate.
GOLDEN_VENDORS = {"nike_ho26", "adidas_premium", "converse_ho26"}
HOLDOUT_VENDORS = {"mizuno_ss27", "hoka_sp27"}


def sample_rows_evenly(rows: list[tuple[int, str]], n: int = 20) -> list[tuple[int, str]]:
    """Pick N rows distributed evenly through the list (deterministic)."""
    if len(rows) <= n:
        return rows
    step = len(rows) / n
    return [rows[int(i * step)] for i in range(n)]


def scaffold(vendor_key: str, n_samples: int = 20) -> Path:
    if vendor_key not in KNOWN_VENDORS:
        raise SystemExit(f"unknown vendor: {vendor_key}. known: {list(KNOWN_VENDORS)}")
    xlsx_name, recon_rel = KNOWN_VENDORS[vendor_key]
    xlsx_path = REPO / xlsx_name
    pdf_glob = list((REPO / f"files/{vendor_key}").glob("*.pdf"))
    pdf_path = pdf_glob[0] if pdf_glob else None
    if not xlsx_path.exists():
        raise SystemExit(f"xlsx not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = []
    for r in range(10, ws.max_row + 1):
        sku = ws.cell(r, 2).value
        if sku:
            rows.append((r, sku))
    sampled = sample_rows_evenly(rows, n_samples)
    skus = []
    for r, sku in sampled:
        entry = {"_verified": False, "_xlsx_row": r}
        for col_idx, field_name in COL_TO_FIELD.items():
            entry[field_name] = ws.cell(r, col_idx).value
        # Add page if it's discoverable (later — leave for human)
        entry["page"] = None
        entry["brand"] = None
        entry["has_photo"] = ws.cell(r, 1).value is not None
        skus.append(entry)

    out_dir = GOLDEN_DIR if vendor_key in GOLDEN_VENDORS else HOLDOUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{vendor_key}.json"
    payload = {
        "vendor_key": vendor_key,
        "pdf_path": str(pdf_path.relative_to(REPO)) if pdf_path else None,
        "description": (
            f"Scaffold from {xlsx_name}. "
            f"Each SKU has `_verified: false` — flip to true after manual verification."
        ),
        "verified_at": None,
        "verified_by": None,
        "set_type": "holdout" if vendor_key in HOLDOUT_VENDORS else "golden",
        "skus": skus,
    }
    out_path.write_text(json.dumps(payload, indent=2, default=str))
    print(f"wrote {out_path}  ({len(skus)} SKUs sampled from {len(rows)} total)")
    return out_path


def main() -> None:
    if len(sys.argv) < 2:
        for v in sorted(KNOWN_VENDORS):
            try:
                scaffold(v)
            except SystemExit as e:
                print(f"  [skip {v}] {e}", file=sys.stderr)
        return
    for v in sys.argv[1:]:
        scaffold(v)


if __name__ == "__main__":
    main()
