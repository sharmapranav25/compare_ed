"""Workbook generation: CatalogExtraction -> BUYSHEET_<vendor>.xlsx.

Opens BUYSHEET_template.xlsx, creates one tab per distinct brand from the
extracted cards (single brand -> rename TEMPLATE tab to vendor name), writes
8 columns + photo per row, embeds photos cropped from card_bbox.

Cells with confidence <0.9 get amber background + cell comment showing the
raw value, page, and source.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl
import pypdfium2 as pdfium
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from PIL import Image as PILImage

from buysheet_v2.consistency import normalize_consistency
from buysheet_v2.lifted.photo_embed import embed_photo
from buysheet_v2.lifted.vocab_normalize import normalize_season, normalize_vendor
from buysheet_v2.schemas.card import ProductCard
from buysheet_v2.schemas.extraction_result import CardConfidence, CatalogExtraction

# photo_vlm and phototune are imported lazily inside the photo-bbox block
# below. They're optional — present in dev (where image-binding experiments
# live) but absent from the shipped release, which runs text-only by default.
# Keeping these out of the top-level import lets the text-only path work in
# either codebase without conditional handling at the call site.

REPO = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = REPO / "BUYSHEET_template.xlsx"

# Column map: ProductCard field name -> 1-indexed Excel column
COL = {
    "photo":            1,   # A
    "sku":              2,   # B  STYLE #
    "mg":               3,   # C  MG
    "sg":               4,   # D  SG
    "ssg":              5,   # E  SSG
    "description":      6,   # F  Item Description
    "color":            7,   # G  Color Desc
    "standard_color":   8,   # H  Standard Color
    "intro_date":      16,   # P  INTRO DATE
    "usd_cost":        22,   # V  USD Cost
    "usd_retail":      23,   # W  USD Retail
}
DATA_ROW_START = 10  # template's first SKU row

AMBER_FILL = PatternFill(start_color="FFE5A8", end_color="FFE5A8", fill_type="solid")
RED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

# openpyxl rejects ASCII control characters in cell values per the OOXML
# spec. PyMuPDF / pdfium occasionally leak form-feed (0x0C) or NULL bytes
# into extracted text when the PDF uses unusual font encodings (Mizuno
# SS27 page 71: "WAVE MG4 LS \x0cffcSUEDE\x0cffcooter"). Strip the entire
# class before any value reaches the xlsx so the workbook write doesn't
# fail an hour into a 90-page extraction. Tab/newline/CR are preserved as
# legitimate whitespace.
import re as _re
_XML_ILLEGAL_RE = _re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _xml_safe(value):
    """Strip ASCII control characters from string values; pass-through otherwise."""
    if isinstance(value, str):
        return _XML_ILLEGAL_RE.sub("", value)
    return value
LOW_CONFIDENCE_THRESHOLD = 0.9      # < this -> amber (review recommended)
CONTRADICTED_THRESHOLD = 0.05       # <= this -> blank + red (oracle says source contradicts)

# Human-readable explanations for confidence/source codes shown in cell comments.
_SOURCE_DESCRIPTIONS = {
    "oracle_verified":              "value appears in the source page text",
    "oracle_verified_relaxed":      "value appears in the source page text (whitespace-relaxed)",
    "vocab_lookup":                 "derived from cached vocabulary (high reliability)",
    "vocab_lookup_confirmed":       "vocabulary lookup, cross-checked against source",
    "cache_vlm_both_valid":         "cached vocab and VLM disagreed; both are plausible",
    "vlm_only_no_region":           "VLM extracted; source text not available to verify",
    "vlm_only_no_source":           "VLM extracted; no source confirmation",
    "vlm_extraction_error_color_eq_description":
                                    "VLM put the description text in the color field — likely wrong",
    "contradicted":                 "the value the model extracted does NOT appear in the source page text",
}


def _comment_for(field_name: str, raw_value, conf_value: float, source: str,
                 page: int, blanked: bool) -> str:
    explanation = _SOURCE_DESCRIPTIONS.get(source, source)
    if blanked:
        raw_repr = f'"{raw_value}"' if raw_value is not None else "(none)"
        return (
            f"AUTO-OMITTED — manual review required\n"
            f"Page {page} · confidence {conf_value:.2f}\n"
            f"Model extracted: {raw_repr}\n"
            f"Reason: {explanation}\n\n"
            f"Open the source PDF at page {page} and fill this cell in by hand."
        )
    return (
        f"Page {page} · confidence {conf_value:.2f}\n"
        f"Source: {explanation}"
    )


def _crop_card_png(pdf: pdfium.PdfDocument, page_no_1: int, bbox_px: list[int],
                   page_width_px: int, page_height_px: int) -> Optional[PILImage.Image]:
    """Crop a card region from the source PDF page render."""
    page = pdf[page_no_1 - 1]
    # Render the full page at the same long-edge as the VLM saw it, so bbox_px
    # coordinates are in the right space.
    long_pt = max(page.get_width(), page.get_height())
    scale = max(page_width_px, page_height_px) / long_pt
    bitmap = page.render(scale=scale)
    pil = bitmap.to_pil()
    x1, y1, x2, y2 = bbox_px
    # Clamp to image bounds
    x1 = max(0, min(pil.size[0], x1))
    y1 = max(0, min(pil.size[1], y1))
    x2 = max(0, min(pil.size[0], x2))
    y2 = max(0, min(pil.size[1], y2))
    if x2 - x1 < 10 or y2 - y1 < 10:
        return None
    return pil.crop((x1, y1, x2, y2)).convert("RGB")


def _confidence_lookup(confidence: list[CardConfidence]) -> dict[tuple[str, int], CardConfidence]:
    return {(c.sku, c.page): c for c in confidence}


def _safe_tab_name(name: str) -> str:
    """Excel worksheet names: max 31 chars, no /\\?*[]:"""
    cleaned = name
    for c in r'/\?*[]:':
        cleaned = cleaned.replace(c, "_")
    return cleaned[:31]


def _write_cell(ws, row: int, col: int, value, comment: Optional[str] = None,
                tier: str = "ok") -> None:
    """Write a cell, applying tier formatting.

    tier: 'ok' (no fill), 'amber' (review recommended), 'red' (auto-omitted).
    For 'red', `value` is ignored and the cell is left blank.
    """
    cell = ws.cell(row, col)
    cell.value = None if tier == "red" else _xml_safe(value)
    if tier == "amber":
        cell.fill = AMBER_FILL
    elif tier == "red":
        cell.fill = RED_FILL
    if comment:
        cell.comment = Comment(_xml_safe(comment), "Kith Buysheet Agent v2")


def _resolve_all_photo_bboxes(
    pdf_path: Path, extraction: CatalogExtraction, page_dims: dict
) -> dict[tuple[int, str], tuple[int, int, int, int]]:
    """Pre-compute per-card photo bboxes for the whole catalog (one PyMuPDF pass).

    Returns {(page, sku): photo_bbox_px} keyed by (page, sku) so duplicate SKUs
    across pages don't collide.
    """
    from buysheet_v2.phototune import resolve_page_photo_bboxes  # noqa: E402
    layout_type = extraction.layout.layout_type if extraction.layout else "grid"
    out: dict[tuple[int, str], tuple[int, int, int, int]] = {}
    for pe in extraction.pages:
        dims = page_dims.get(pe.page)
        if not dims or not pe.cards:
            continue
        w, h = dims
        page_resolved = resolve_page_photo_bboxes(
            pdf_path, pe.page, pe.cards, layout_type, w, h,
        )
        for sku, bbox in page_resolved.items():
            out[(pe.page, sku)] = bbox
    return out


def write_workbook(
    extraction: CatalogExtraction,
    out_path: Path,
    *,
    pdf_path: Optional[Path] = None,
    catalog_brand_hint: Optional[str] = None,
    embed_photos: bool = True,
) -> Path:
    """Write the extraction to a BUYSHEET-shaped xlsx.

    Multi-brand handling is delegated to consistency.normalize_consistency,
    which votes brand by SKU prefix family + partitions cards. When ≥2
    brands are present, one worksheet per brand is created; otherwise the
    single TEMPLATE sheet is used.

    pdf_path: optional path to the source PDF. When provided, photo bboxes
    are cropped from the original page renders and embedded in col A.
    catalog_brand_hint: fallback brand for cards with brand=None (used in
    single-brand catalogs where the brand isn't repeated on every card).
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"template missing: {TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    template_ws = wb["TEMPLATE"]
    conf_lookup = _confidence_lookup(extraction.confidence)

    all_cards = extraction.all_cards
    cons = normalize_consistency(all_cards, catalog_brand=catalog_brand_hint)
    partitions = cons["partitions"]
    multi_brand = cons["is_multi_brand"]
    distinct_brands = sorted(b for b in partitions if b and b != "_unbranded")

    pdf = None
    if embed_photos and pdf_path is not None:
        try:
            pdf = pdfium.PdfDocument(str(pdf_path))
        except Exception:
            pdf = None
    # Map page_no -> (width_px, height_px) for crop scaling
    page_dims: dict[int, Optional[tuple[int, int]]] = {pe.page: None for pe in extraction.pages}
    if pdf is not None:
        from buysheet_v2.lifted.pdf_render import VLM_MAX_LONG_EDGE_PX
        for page_no in page_dims:
            page = pdf[page_no - 1]
            long_pt = max(page.get_width(), page.get_height())
            scale = VLM_MAX_LONG_EDGE_PX / long_pt
            w = int(round(page.get_width() * scale))
            h = int(round(page.get_height() * scale))
            page_dims[page_no] = (w, h)

    # Photo-bbox resolution (two-stage), only when embed_photos is set:
    #   1. photo_vlm.resolve_catalog_photo_bboxes — per-card VLM extraction
    #      with SKU-marker annotation. Sidecar-cached so re-runs of write.py
    #      cost $0. ~$0.003/SKU when fresh.
    #   2. phototune.resolve_page_photo_bboxes — deterministic heuristic
    #      (PyMuPDF SKU anchor + row-context geometry). Used as fallback for
    #      any SKU the per-card VLM couldn't resolve (image-only pages,
    #      transient VLM failures).
    # When embed_photos=False (Slack-bot default), col A is left empty and the
    # workbook ships as the validated text-only v1 deliverable.
    resolved_photo: dict[tuple[int, str], tuple[int, int, int, int]] = {}
    if embed_photos and pdf_path is not None:
        from buysheet_v2.photo_vlm import resolve_catalog_photo_bboxes  # noqa: E402
        try:
            vlm_resolved, vlm_usage = resolve_catalog_photo_bboxes(
                pdf_path, extraction, page_dims,
            )
            resolved_photo.update(vlm_resolved)
            print(f"[write] photo_vlm resolved {len(vlm_resolved)} bboxes  "
                  f"({vlm_usage['calls']} fresh VLM calls)")
        except Exception as e:
            print(f"[write] photo_vlm failed: {type(e).__name__}: {e}; "
                  f"falling back to heuristic only")
        # Heuristic fallback for SKUs the VLM didn't resolve
        heuristic = _resolve_all_photo_bboxes(pdf_path, extraction, page_dims)
        for key, bbox in heuristic.items():
            resolved_photo.setdefault(key, bbox)

    season = normalize_season(extraction.vendor_key, wb)

    if multi_brand:
        ws_map: dict[str, openpyxl.worksheet.worksheet.Worksheet] = {}
        # Rename the template tab to the first brand; clone for subsequent brands
        first = distinct_brands[0]
        template_ws.title = _safe_tab_name(first)
        ws_map[first] = template_ws
        for brand in distinct_brands[1:]:
            ws_map[brand] = wb.copy_worksheet(template_ws)
            ws_map[brand].title = _safe_tab_name(brand)
        # _unbranded cards (no brand) get appended to the dominant tab
        unbranded = partitions.get("_unbranded", [])
        if unbranded:
            partitions[first] = partitions.get(first, []) + unbranded
        for brand_name, ws in ws_map.items():
            _populate_workbook(
                ws, partitions.get(brand_name, []), conf_lookup,
                pdf, page_dims, brand_name, resolved_photo,
            )
            vendor_match = normalize_vendor(brand_name, wb) or brand_name
            ws["B1"] = vendor_match
            if season:
                ws["B2"] = season
    else:
        # Single-brand case: keep the TEMPLATE sheet, set B1
        target_brand = distinct_brands[0] if distinct_brands else None
        _populate_workbook(
            template_ws, all_cards, conf_lookup,
            pdf, page_dims, target_brand, resolved_photo,
        )
        if target_brand:
            vendor_match = normalize_vendor(target_brand, wb) or target_brand
            template_ws["B1"] = vendor_match
        if season:
            template_ws["B2"] = season

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    if pdf is not None:
        pdf.close()
    return out_path


def _populate_workbook(
    ws,
    cards: list[ProductCard],
    conf_lookup: dict[tuple[str, int], CardConfidence],
    pdf: Optional[pdfium.PdfDocument],
    page_dims: dict[int, Optional[tuple[int, int]]],
    brand_name: Optional[str],
    resolved_photo: dict[tuple[int, str], tuple[int, int, int, int]],
) -> None:
    """Write a card list to one worksheet starting at DATA_ROW_START."""
    for i, card in enumerate(cards):
        row = DATA_ROW_START + i
        if row > 883:
            break  # template hard cap
        conf = conf_lookup.get((card.sku, card.page))

        for field_name, col in COL.items():
            if field_name == "photo":
                continue
            if field_name == "sku":
                # SKU is the row key — always written, never blanked.
                value = card.sku
                conf_value = 1.0
                source = "sku_anchor"
            else:
                value = getattr(card, field_name, None)
                if value is None:
                    continue
                conf_value = (conf.per_field.get(field_name) if conf else None) or 0.0
                source = (conf.per_field_source.get(field_name) if conf else None) or "vlm"

            if field_name != "sku" and conf_value <= CONTRADICTED_THRESHOLD:
                tier = "red"
            elif field_name != "sku" and conf_value < LOW_CONFIDENCE_THRESHOLD:
                tier = "amber"
            else:
                tier = "ok"

            comment = _comment_for(
                field_name, value, conf_value, source,
                page=card.page, blanked=(tier == "red"),
            )
            _write_cell(ws, row, col, value, comment=comment, tier=tier)

        # Photo: use the deterministic bbox from phototune (PyMuPDF text-search
        # + layout-aware heuristic, with VLM bbox reconciled in when it sits
        # inside the deterministic region). Falls back to VLM bbox alone on
        # image-only pages where text-search returns nothing. Padded to avoid
        # silhouette cut-offs.
        if pdf is not None and page_dims.get(card.page):
            w, h = page_dims[card.page]
            tuned = resolved_photo.get((card.page, card.sku))
            crop_bbox = tuned or card.photo_bbox_px or card.card_bbox_px
            if tuned:
                crop_source = "phototune"
            elif card.photo_bbox_px:
                crop_source = "vlm_photo_bbox"
            else:
                crop_source = "card_bbox"
            if crop_bbox:
                img = _crop_card_png(pdf, card.page, list(crop_bbox), w, h)
                if img is not None:
                    embed_photo(
                        ws, row, COL["photo"], img,
                        comment_text=(
                            f"sku={card.sku}  page={card.page}  brand={brand_name or 'n/a'}  "
                            f"crop_source={crop_source}  bbox={list(crop_bbox)}"
                        ),
                    )
