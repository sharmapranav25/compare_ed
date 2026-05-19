"""Excel adapter — deterministic, no LLM calls.

Reads an .xlsx with openpyxl, finds a header row by matching column
names against a synonym table, and emits <xlsx-stem>.pages/00.json in
the same shape extract_products writes. Every emitted product carries
verification: "deterministic" on each field so the existing verifier-
aware build code treats them as trusted-without-VLM.

Vendor is inferred from the filename slug (e.g. nike_pricing.xlsx → NIKE);
override via the `vendor=` kwarg.

In milestone 1 the header mapping is purely synonym-table based. If the
required fields (sku, description) can't be mapped, the sheet is skipped
with an error recorded in the page JSON — no LLM fallback yet.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

# Header synonyms → canonical product field. Lowercased + normalized
# (non-alphanumeric → underscore) before lookup. Conservative on purpose:
# we'd rather skip a sheet with "Item Cost (USD)" than mis-map "Cost Code".
SYNONYMS: dict[str, str] = {
    # SKU
    "sku": "sku", "skus": "sku", "style": "sku", "style_no": "sku",
    "style_number": "sku", "style_code": "sku", "item": "sku",
    "item_no": "sku", "item_number": "sku", "item_code": "sku",
    "model": "sku", "model_no": "sku", "model_number": "sku",
    "article": "sku", "article_no": "sku", "product_code": "sku",
    "upc": "sku",
    # description
    "description": "description", "desc": "description",
    "item_description": "description", "product_name": "description",
    "style_name": "description", "name": "description",
    "product_description": "description",
    # color
    "color": "color", "colour": "color", "colorway": "color",
    "color_name": "color", "color_description": "color",
    "color_desc": "color", "colour_name": "color",
    # cost (wholesale)
    "cost": "cost", "wholesale": "cost", "wholesale_cost": "cost",
    "wholesale_price": "cost", "whsl": "cost", "whsle": "cost",
    "wsp": "cost", "unit_cost": "cost", "trade_price": "cost",
    "usd_cost": "cost", "wholesale_usd": "cost",
    # retail
    "retail": "retail", "msrp": "retail", "srp": "retail", "rrp": "retail",
    "retail_price": "retail", "selling_price": "retail",
    "suggested_retail": "retail", "usd_retail": "retail",
    # intro / drop date
    "intro": "intro_date", "intro_date": "intro_date",
    "release": "intro_date", "release_date": "intro_date",
    "launch": "intro_date", "launch_date": "intro_date",
    "instore": "intro_date", "in_store": "intro_date",
    "in_store_date": "intro_date", "drop": "intro_date",
    "drop_date": "intro_date", "available": "intro_date",
    "available_date": "intro_date", "ship": "intro_date",
    "ship_date": "intro_date", "delivery": "intro_date",
    "delivery_date": "intro_date",
    # gender / department
    "gender": "gender_hint", "department": "gender_hint",
    "dept": "gender_hint", "category": "gender_hint",
    "division": "gender_hint",
}

REQUIRED_FIELDS = ("sku", "description")
PRODUCT_FIELDS = ("sku", "description", "color", "cost",
                  "retail", "intro_date", "gender_hint")
HEADER_SCAN_ROWS = 20
NON_SKU_PATTERNS = re.compile(
    r"(?i)^\s*(total|subtotal|grand[\s_-]*total|sum)\s*$"
)


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def _slug_vendor_from_filename(stem: str) -> str:
    """nike_HO26 → NIKE, nike-pricing-sheet → NIKE."""
    parts = re.split(r"[_\-\s]+", stem)
    parts = [p for p in parts if p]
    return parts[0].upper() if parts else stem.upper()


def _stringify(v) -> str | None:
    """Excel cell → string, or None for empty. Floats/ints/dates preserve
    enough info for downstream parse_money / vocab_map to handle them."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def _find_header_row(ws) -> tuple[int, dict[int, str]] | None:
    """Scan up to HEADER_SCAN_ROWS for the row with the most synonym hits.
    Returns (1-based row index, {1-based col index: canonical field}) or
    None if no row has ≥3 hits AND both required fields mapped."""
    best: tuple[int, dict[int, str]] | None = None
    for r in range(1, min(HEADER_SCAN_ROWS, ws.max_row or 0) + 1):
        col_map: dict[int, str] = {}
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            field = SYNONYMS.get(_norm_header(str(v)))
            if field and c not in col_map and field not in col_map.values():
                col_map[c] = field
        n_hits = len(col_map)
        if n_hits < 3:
            continue
        if not all(f in col_map.values() for f in REQUIRED_FIELDS):
            continue
        if best is None or n_hits > len(best[1]):
            best = (r, col_map)
    return best


def _row_to_product(row_values: list, col_map: dict[int, str]) -> dict | None:
    """Build a product dict from one data row. Returns None if the row
    lacks a SKU or looks like a totals/section line."""
    p: dict = {f: None for f in PRODUCT_FIELDS}
    for col, field in col_map.items():
        if col - 1 < len(row_values):
            p[field] = _stringify(row_values[col - 1])
    sku = p.get("sku")
    if not sku:
        return None
    if NON_SKU_PATTERNS.match(sku):
        return None
    # Tag every field as deterministic so the existing
    # _VERIFICATION_LABEL audit treats them as trusted-without-VLM.
    p["verification"] = {f: "deterministic" for f in
                         ("sku", "description", "color", "cost",
                          "retail", "intro_date")}
    return p


def normalize(xlsx_path: Path, *, vendor: str | None = None,
              force: bool = False) -> Path:
    """Read xlsx → write <stem>.pages/00.json. Returns pages_dir."""
    pages_dir = xlsx_path.with_name(xlsx_path.stem + ".pages")
    pages_dir.mkdir(exist_ok=True)
    out_json = pages_dir / "00.json"
    if out_json.exists() and not force:
        return pages_dir

    wb = load_workbook(xlsx_path, data_only=True)
    inferred_vendor = vendor or _slug_vendor_from_filename(xlsx_path.stem)
    products: list[dict] = []
    sheet_errors: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        found = _find_header_row(ws)
        if found is None:
            sheet_errors.append(f"{sheet_name}: no header row with ≥3 known fields")
            continue
        header_row, col_map = found
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            p = _row_to_product(list(row), col_map)
            if p is not None:
                products.append(p)
    wb.close()

    record: dict = {
        "page_no": 0,
        "label": "product",
        "source_kind": "excel",
        "prev_context": {"vendor": inferred_vendor, "current_section": None},
        "context_after": {"vendor": inferred_vendor, "current_section": None},
        "products": products,
        "rejected_candidates": [],
    }
    # Excel rows are deterministic — text-layer verification doesn't apply.
    # We set this explicitly so the build code's flag heuristics (which
    # otherwise treat absence-of-key as "PDF, no info") behave sensibly.
    record["text_layer_present"] = True
    if not products and sheet_errors:
        record["error"] = "; ".join(sheet_errors)
    elif sheet_errors:
        record["sheet_errors"] = sheet_errors
    out_json.write_text(json.dumps(record, indent=2))
    return pages_dir
