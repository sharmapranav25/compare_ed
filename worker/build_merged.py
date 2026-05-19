"""Write a merged buy-sheet xlsx from pre-merged products_rows.

Reuses the dropdown-resolution and row-writing helpers from
build_buysheet.py without modifying it. The merge step itself lives in
worker/merge.py; this module owns only the build half: template load,
vocab_map dispatch, row write, REVIEW sheet, SKU_CONFLICTS sheet, save.

The REVIEW sheet is local to this module so we can include a `source`
column and tolerate non-int page_no values (they're never produced by
the single-doc path, but the multi-doc one has them).

The SKU_CONFLICTS sheet surfaces SKUs that the merge step refused to
collapse (because the source pages disagreed on some non-empty field);
those rows appear in TEMPLATE with yellow STYLE# tint and one row per
source occurrence, and the conflict detail (which fields, which values)
lives in this sheet so a reviewer can resolve them by hand.
"""

from __future__ import annotations

import sys
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from build_buysheet import (  # noqa: E402
    REVIEW_FILL,
    REVIEW_SHEET_NAME,
    TEMPLATE_PATH,
    gather_unique_tasks,
    load_dropdown_vocabs,
    resolve_all,
    write_rows,
)
from vocab_map import load_cache, save_cache  # noqa: E402

load_dotenv()

CONFLICTS_SHEET_NAME = "SKU_CONFLICTS"


def _write_review_sheet(wb, page_records: list[dict]) -> None:
    if REVIEW_SHEET_NAME in wb.sheetnames:
        del wb[REVIEW_SHEET_NAME]
    flagged = [r for r in page_records if r.get("flagged")]
    ws = wb.create_sheet(REVIEW_SHEET_NAME)
    ws.append(["source", "page_no", "label", "n_products", "error"])
    for c in ws[1]:
        c.fill = REVIEW_FILL
    for r in flagged:
        ws.append([
            r.get("source", ""),
            str(r.get("page_no", "")),
            r.get("label", ""),
            r.get("n_products", 0),
            r.get("error") or "",
        ])
    if not flagged:
        ws.append(["(no flagged pages)", "", "", "", ""])


def _write_conflicts_sheet(wb, sku_conflicts: list[dict]) -> None:
    """One row per (SKU, occurrence). For a SKU that conflicts across two
    sources you get two rows here, plus a header showing which fields
    disagreed."""
    if CONFLICTS_SHEET_NAME in wb.sheetnames:
        del wb[CONFLICTS_SHEET_NAME]
    ws = wb.create_sheet(CONFLICTS_SHEET_NAME)
    ws.append(["sku", "conflicting_fields", "source", "page_no",
               "sku_raw", "value (per conflicting field)"])
    for c in ws[1]:
        c.fill = REVIEW_FILL
    if not sku_conflicts:
        ws.append(["(no SKU conflicts)", "", "", "", "", ""])
        return
    for conflict in sku_conflicts:
        sku = conflict["sku_canon"]
        fields = ", ".join(conflict["conflicting_fields"])
        for occ in conflict["occurrences"]:
            value_str = "; ".join(
                f"{f}={occ['values'].get(f)!r}"
                for f in conflict["conflicting_fields"]
            )
            ws.append([
                sku,
                fields,
                occ.get("source", ""),
                str(occ.get("page_no", "")),
                occ.get("sku_raw", ""),
                value_str,
            ])


def build_merged(products_rows: list[dict], page_records: list[dict],
                 sku_conflicts: list[dict], vendor: str | None,
                 out_path: Path, workers: int) -> Path:
    if not products_rows:
        print("no products to write; skipping", file=sys.stderr)
        return out_path

    vocabs = load_dropdown_vocabs(TEMPLATE_PATH)
    caches = {field: load_cache(field) for field in vocabs}
    client = anthropic.Anthropic(max_retries=10)

    tasks = gather_unique_tasks(products_rows, vocabs)
    resolutions = resolve_all(tasks, caches, client, workers)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["TEMPLATE"]
    if vendor:
        ws["B1"] = vendor

    print(f"writing {len(products_rows)} merged product rows", file=sys.stderr)
    write_rows(ws, products_rows, resolutions)
    _write_review_sheet(wb, page_records)
    _write_conflicts_sheet(wb, sku_conflicts)

    for field, cache in caches.items():
        save_cache(field, cache)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    n_flagged_pages = sum(1 for r in page_records if r.get("flagged"))
    n_conflict_skus = len(sku_conflicts)
    n_conflict_rows = sum(len(c["occurrences"]) for c in sku_conflicts)
    print("", file=sys.stderr)
    print(f"SUMMARY: {len(products_rows)} merged rows, "
          f"{n_flagged_pages} flagged source pages, "
          f"{n_conflict_skus} conflicting SKUs ({n_conflict_rows} rows kept) "
          f"-> {out_path}", file=sys.stderr)
    return out_path
