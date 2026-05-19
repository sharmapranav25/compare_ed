"""Step 3 of agentic per-page parsing: aggregate per-page product extracts
into BUYSHEET_<vendor>.xlsx.

Reads every <doc>.pages/<NN>.json produced by classify_pages.py +
extract_products.py, walks pages in order, and writes one row per product to
a copy of BUYSHEET_template.xlsx (data starts at row 10).

Per-cell mapping policy:
  - Free-text columns (STYLE #, Item Description, Color Desc, USD Cost,
    USD Retail) → raw vendor value verbatim. Money fields are parsed into
    numbers when possible; non-numeric values are written as raw text.
  - Dropdown columns (MG, SG, SSG, Standard Color, INTRO DATE) → vocab_map
    asks Sonnet to pick a canonical from the closed list. Hit → write
    canonical (orange-filled to mark LLM-resolved). Miss but vendor value
    present → write raw vendor value (overrides the dropdown). Unsure /
    absent → leave the cell empty.

Failure surfacing:
  - Every page JSON with `error: "<reason>"` is flagged.
  - Pages classified as `product` but that extracted 0 products are flagged.
  - For each product on a flagged page, the STYLE# cell is tinted yellow.
  - A REVIEW sheet is appended to the xlsx listing every flagged page.

Concurrency:
  - The vocab_map.map_to_dropdown LLM calls (Sonnet) are dispatched via
    a ThreadPoolExecutor; identical (field, key) tuples are deduped before
    submission so each unique vendor-value is resolved exactly once.
  - openpyxl writes happen serially after all resolutions complete.

Usage:
  python agentic_parsing/build_buysheet.py files/<vendor>/<doc>.pdf
  python agentic_parsing/build_buysheet.py files/<vendor>/<doc>.pdf --out path.xlsx
  python agentic_parsing/build_buysheet.py files/<vendor>/<doc>.pdf --workers 8
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import PatternFill
from openpyxl.utils.units import pixels_to_EMU

sys.path.insert(0, str(Path(__file__).resolve().parent))
from analysis.usage import add_usage, empty_usage  # noqa: E402
from vocab_map import load_cache, map_to_dropdown, save_cache  # noqa: E402

load_dotenv()

REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = REPO_ROOT / "BUYSHEET_template.xlsx"
DATA_START_ROW = 10
HEADER_ROW = 9
AI_FILL = PatternFill(start_color="FFD7A8", end_color="FFD7A8", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
REVIEW_SHEET_NAME = "REVIEW"

# buy-sheet column letters for the fields we populate
COL = {
    "image":           "A",
    "sku":             "B",
    "mg":              "C",
    "sg":              "D",
    "ssg":             "E",
    "description":     "F",
    "color":           "G",
    "standard_color":  "H",
    "intro_date":      "P",
    "cost":            "V",
    "retail":          "W",
}

# Image-column geometry. row height is in points (openpyxl convention);
# col width is in "characters" (openpyxl convention, ~7px per unit).
IMAGE_TARGET_PX     = 128
IMAGE_ROW_HEIGHT_PT = 100
IMAGE_COL_WIDTH     = 20

# which xlsx field names use which dropdown column in Product Data
DROPDOWN_FIELDS = {
    "MG":             {"col": 1, "buyseet_col": "mg"},
    "SG":             {"col": 2, "buyseet_col": "sg"},
    "SSG":            {"col": 3, "buyseet_col": "ssg"},
    "Standard Color": {"col": 4, "buyseet_col": "standard_color"},
    "INTRO DATE":     {"col": 7, "buyseet_col": "intro_date"},
}

MONEY_RE = re.compile(r"(\d{1,3}(?:[,\d]{0,6})(?:\.\d{1,2})?)")


def load_dropdown_vocabs(template: Path) -> dict[str, list[str]]:
    wb = load_workbook(template, read_only=True, data_only=True)
    pd = wb["Product Data"]
    vocabs: dict[str, list[str]] = {}
    for field, spec in DROPDOWN_FIELDS.items():
        col = spec["col"]
        values: list[str] = []
        for row in range(2, pd.max_row + 1):
            v = pd.cell(row=row, column=col).value
            if v is None:
                continue
            values.append(str(v).strip())
        vocabs[field] = values
    wb.close()
    return vocabs


_VERIFICATION_LABEL = {
    "sku":         ("SKUs dropped",       "deterministic"),
    "cost":        ("costs cleared",      "not_in_text_layer"),
    "retail":      ("retails cleared",    "not_in_text_layer"),
    "description": ("descriptions unverified", "unverified"),
    "color":       ("colors unverified",  "unverified"),
    "intro_date":  ("intro dates unverified", "unverified"),
}


def _verification_summary(rec: dict) -> str:
    """Compact 'N x, M y' summary of verification issues on one page."""
    parts: list[str] = []
    if rec.get("text_layer_present") is False:
        cause = rec.get("text_layer_error") or "text_layer_absent"
        parts.append(cause)
    rejected = rec.get("rejected_candidates") or []
    n_det = sum(1 for r in rejected if r.get("stage") == "deterministic")
    if n_det:
        parts.append(f"{n_det} {_VERIFICATION_LABEL['sku'][0]}")
    products = rec.get("products") or []
    for field, (label, expected) in _VERIFICATION_LABEL.items():
        if field == "sku":
            continue
        n = sum(1 for p in products
                if (p.get("verification") or {}).get(field) == expected)
        if n:
            parts.append(f"{n} {label}")
    return "; ".join(parts)


def collect_pages(pdf_path: Path) -> tuple[list[dict], list[dict], str | None]:
    """Walk <doc>.pages/*.json in page order. Returns:
       (products_rows, page_records, vendor)

    products_rows: [{page_no, context, product, page_flagged}]
    page_records:  [{page_no, label, error, n_products, flagged,
                     verification_issues}] for every page
    vendor: first non-empty vendor seen in any context_after
    """
    pages_dir = pdf_path.with_name(pdf_path.stem + ".pages")
    if not pages_dir.exists():
        print(f"error: {pages_dir} not found — run classify_pages.py + "
              f"extract_products.py first", file=sys.stderr)
        sys.exit(1)
    products_rows: list[dict] = []
    page_records: list[dict] = []
    vendor: str | None = None
    for jf in sorted(pages_dir.glob("*.json")):
        rec = json.loads(jf.read_text())
        page_no = rec.get("page_no")
        label = rec.get("label", "unknown")
        error = rec.get("error")
        ctx = rec.get("context_after") or rec.get("prev_context") or {}
        if ctx.get("vendor") and not vendor:
            vendor = ctx["vendor"]
        products = rec.get("products") or []
        verification_issues = _verification_summary(rec)
        # flag: page errored, OR page was classified as product but no
        # products extracted, OR PyMuPDF couldn't read the text layer, OR
        # any verification marker is non-"ok". Helps the reviewer find
        # dropouts and unverified fields.
        flagged = (bool(error)
                   or (label == "product" and not products)
                   or rec.get("text_layer_present") is False
                   or bool(verification_issues))
        page_records.append({
            "page_no": page_no, "label": label, "error": error,
            "n_products": len(products), "flagged": flagged,
            "verification_issues": verification_issues,
        })
        for p in products:
            products_rows.append({
                "page_no": page_no,
                "context": ctx,
                "product": p,
                "page_flagged": flagged,
            })
    return products_rows, page_records, vendor


def parse_money(text):
    if text is None:
        return None, None
    if isinstance(text, (int, float)):
        return float(text), str(text)
    s = str(text).strip()
    if not s:
        return None, None
    m = MONEY_RE.search(s.replace(",", ""))
    if m:
        try:
            return float(m.group(1).replace(",", "")), s
        except ValueError:
            return None, s
    return None, s


def _signals_for(field: str, product: dict, ctx: dict) -> tuple[str, dict, str | None]:
    """Returns (cache_key, signals_dict, raw_fallback) for a given field."""
    p = product
    sku = (p.get("sku") or "").strip()
    desc = (p.get("description") or "").strip()
    color = (p.get("color") or "").strip()
    intro = (p.get("intro_date") or "").strip()
    gender_hint = (p.get("gender_hint") or "").strip()
    section = (ctx.get("current_section") or "").strip()
    if field == "MG":
        key = "|".join(filter(None, [desc, section, gender_hint])) or sku
        signals = {"description": desc, "section": section,
                   "sku": sku, "gender_hint": gender_hint}
        return key, signals, (gender_hint or section or None)
    if field == "SG":
        key = desc or section
        return key, {"description": desc, "section": section}, (section or None)
    if field == "SSG":
        key = desc or section
        return key, {"description": desc, "section": section}, None
    if field == "Standard Color":
        return color, {"color": color, "description": desc}, (color or None)
    if field == "INTRO DATE":
        return intro, {"date_as_printed": intro}, (intro or None)
    raise ValueError(f"unknown field {field}")


def gather_unique_tasks(products_rows: list[dict],
                        vocabs: dict[str, list[str]]) -> dict[tuple[str, str], dict]:
    """Dedupe (field, key) across all products. Identical lookups share one task."""
    tasks: dict[tuple[str, str], dict] = {}
    for row in products_rows:
        for field in DROPDOWN_FIELDS:
            key, signals, _ = _signals_for(field, row["product"], row["context"])
            if not (key and key.strip()):
                continue
            tk = (field, key.strip().lower())
            if tk in tasks:
                continue
            tasks[tk] = {"field": field, "choices": vocabs[field],
                         "key": key, "signals": signals}
    return tasks


def resolve_all(tasks: dict[tuple[str, str], dict],
                caches: dict[str, dict], client, workers: int) -> dict[tuple[str, str], dict]:
    """Run all unique vocab_map calls in a thread pool. Returns dict of results.
    Each cache (one per field) gets its own lock so two threads don't both
    issue an LLM call for the same key — first thread wins, second hits the
    mutated cache."""
    cache_locks = {f: threading.Lock() for f in caches}
    results: dict[tuple[str, str], dict] = {}

    def _one(task_key, spec):
        field = spec["field"]
        with cache_locks[field]:
            # Re-check cache under the lock — a sibling task with same
            # normalized key may have populated it already.
            res = map_to_dropdown(
                field=field, choices=spec["choices"],
                key=spec["key"], signals=spec["signals"],
                client=client, cache=caches[field],
            )
        return task_key, res

    print(f"resolving {len(tasks)} unique (field, key) lookups with {workers} workers",
          file=sys.stderr)
    n_done = 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futs = {pool.submit(_one, tk, spec): tk for tk, spec in tasks.items()}
        for fut in as_completed(futs):
            tk, res = fut.result()
            results[tk] = res
            n_done += 1
            if n_done % 25 == 0 or n_done == len(tasks):
                print(f"  resolved {n_done}/{len(tasks)}", file=sys.stderr)
    return results


def _write(cell, value, llm_resolved: bool) -> None:
    if value is None or value == "":
        return
    cell.value = value
    if llm_resolved:
        cell.fill = AI_FILL


def _setup_image_column(ws) -> None:
    """Widen column A and ensure a header label at A<HEADER_ROW>.

    The shipped BUYSHEET template already has "PHOTO" at A9; we leave
    existing headers alone and only fill in "PHOTO" as a fallback when
    a custom template lacks one. Width is bumped unconditionally — the
    default 12-char width is too narrow for a 128px thumbnail.
    """
    ws.column_dimensions[COL["image"]].width = IMAGE_COL_WIDTH
    header_cell = ws[f"{COL['image']}{HEADER_ROW}"]
    if not header_cell.value:
        header_cell.value = "PHOTO"


def _embed_image(ws, image_path: Path, excel_row: int) -> None:
    """Anchor an image to a single cell (OneCellAnchor) so it scrolls
    with the row. Sized in EMU; the row height grows to match the image
    so the thumbnail fits visually inside the cell bounds.

    OneCellAnchor pins both top-left corner and absolute size — the
    image stays put when its anchor row moves, but does NOT resize when
    column/row dimensions change (which is what we want for thumbnails).
    """
    img = XLImage(str(image_path))
    img.width = IMAGE_TARGET_PX
    img.height = IMAGE_TARGET_PX
    col_idx = ord(COL["image"]) - ord("A")
    marker = AnchorMarker(
        col=col_idx, colOff=pixels_to_EMU(4),
        row=excel_row - 1, rowOff=pixels_to_EMU(4),
    )
    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(IMAGE_TARGET_PX),
            cy=pixels_to_EMU(IMAGE_TARGET_PX),
        ),
    )
    ws.add_image(img)
    current = ws.row_dimensions[excel_row].height
    if current is None or current < IMAGE_ROW_HEIGHT_PT:
        ws.row_dimensions[excel_row].height = IMAGE_ROW_HEIGHT_PT


def resolved_value(res: dict, raw_fallback: str | None) -> tuple[object, bool]:
    """Three-branch policy (see docstring at top of file)."""
    if res["confidence"] == "low":
        return None, False
    if res["value"] is not None:
        return res["value"], (res["source"] == "llm")
    if raw_fallback and str(raw_fallback).strip():
        return str(raw_fallback).strip(), False
    return None, False


def write_rows(ws, products_rows: list[dict],
               resolutions: dict[tuple[str, str], dict]) -> None:
    for i, row in enumerate(products_rows):
        excel_row = DATA_START_ROW + i
        p = row["product"]
        ctx = row["context"]
        sku = (p.get("sku") or "").strip()
        desc = (p.get("description") or "").strip()
        color = (p.get("color") or "").strip()
        cost_num, cost_raw = parse_money(p.get("cost"))
        retail_num, retail_raw = parse_money(p.get("retail"))

        # Free-text + numeric cells
        sku_cell = ws[f"{COL['sku']}{excel_row}"]
        _write(sku_cell, sku, False)
        if row["page_flagged"] and sku_cell.value is not None:
            sku_cell.fill = REVIEW_FILL
        _write(ws[f"{COL['description']}{excel_row}"], desc, False)
        _write(ws[f"{COL['color']}{excel_row}"], color, False)
        _write(ws[f"{COL['cost']}{excel_row}"],
               cost_num if cost_num is not None else cost_raw, False)
        _write(ws[f"{COL['retail']}{excel_row}"],
               retail_num if retail_num is not None else retail_raw, False)

        # Dropdown cells — look up from resolutions
        for field, col_key in [("MG", "mg"), ("SG", "sg"), ("SSG", "ssg"),
                               ("Standard Color", "standard_color"),
                               ("INTRO DATE", "intro_date")]:
            key, _, raw_fallback = _signals_for(field, p, ctx)
            if not (key and key.strip()):
                continue
            res = resolutions.get((field, key.strip().lower()))
            if res is None:
                continue
            value, llm = resolved_value(res, raw_fallback)
            _write(ws[f"{COL[col_key]}{excel_row}"], value, llm)

        # Image cell — anchored thumbnail. Silently skips when image_path
        # is absent (Excel-source rows or pages without detections).
        img_path = p.get("image_path")
        if img_path and Path(img_path).exists():
            try:
                _embed_image(ws, Path(img_path), excel_row)
            except Exception as exc:  # noqa: BLE001
                print(f"  warn: image embed failed for {img_path}: "
                      f"{type(exc).__name__}: {exc}", file=sys.stderr)

        if (i + 1) % 25 == 0 or (i + 1) == len(products_rows):
            print(f"  wrote row {i + 1}/{len(products_rows)}: {sku!r}",
                  file=sys.stderr)


def write_review_sheet(wb, page_records: list[dict]) -> None:
    """Append (or refresh) a REVIEW sheet listing every page with an issue."""
    if REVIEW_SHEET_NAME in wb.sheetnames:
        del wb[REVIEW_SHEET_NAME]
    flagged = [r for r in page_records if r["flagged"]]
    ws = wb.create_sheet(REVIEW_SHEET_NAME)
    ws.append(["page_no", "label", "n_products", "error", "verification_issues"])
    for c in ws[1]:
        c.fill = REVIEW_FILL
    for r in flagged:
        ws.append([r["page_no"], r["label"], r["n_products"],
                   r["error"] or "", r.get("verification_issues") or ""])
    if not flagged:
        ws.append(["(no flagged pages)", "", "", "", ""])


def build(pdf_path: Path, out_path: Path, workers: int) -> None:
    products_rows, page_records, vendor = collect_pages(pdf_path)
    if not products_rows:
        print("no products extracted; nothing to write", file=sys.stderr)
        return
    vocabs = load_dropdown_vocabs(TEMPLATE_PATH)
    caches = {field: load_cache(field) for field in vocabs}
    client = anthropic.Anthropic(max_retries=10)

    tasks = gather_unique_tasks(products_rows, vocabs)
    resolutions = resolve_all(tasks, caches, client, workers)

    # Aggregate vocab_map usage (each LLM call returns "usage"; cache hits
    # don't carry one). Write a sidecar JSON the analysis step picks up.
    build_usage = empty_usage()
    for res in resolutions.values():
        u = res.get("usage")
        if u:
            add_usage(build_usage, u)
    pages_dir = pdf_path.with_name(pdf_path.stem + ".pages")
    if pages_dir.exists():
        (pages_dir / "_build_usage.json").write_text(
            json.dumps({"vocab_map": build_usage}, indent=2, sort_keys=True)
        )

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["TEMPLATE"]
    if vendor:
        ws["B1"] = vendor

    if any(r["product"].get("image_path") for r in products_rows):
        _setup_image_column(ws)

    print(f"writing {len(products_rows)} product rows", file=sys.stderr)
    write_rows(ws, products_rows, resolutions)
    write_review_sheet(wb, page_records)

    for field, cache in caches.items():
        save_cache(field, cache)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    flagged_count = sum(1 for r in page_records if r["flagged"])
    print("", file=sys.stderr)
    print(f"SUMMARY: {len(products_rows)} rows written, "
          f"{flagged_count} flagged pages -> {out_path}", file=sys.stderr)
    if flagged_count:
        print("Flagged pages (see REVIEW sheet):", file=sys.stderr)
        for r in page_records:
            if not r["flagged"]:
                continue
            if r["error"]:
                detail = r["error"]
            elif r["label"] == "product" and not r["n_products"]:
                detail = f"label={r['label']}, 0 products"
            else:
                detail = r.get("verification_issues") or f"label={r['label']}"
            print(f"  page {r['page_no']:02d}: {detail}", file=sys.stderr)


def _slugify(s: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower() or "vendor"


def default_out_path(pdf_path: Path, vendor_hint: str | None) -> Path:
    slug = _slugify(vendor_hint or pdf_path.parent.name or pdf_path.stem)
    return REPO_ROOT / f"BUYSHEET_{slug}.xlsx"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("pdf", type=Path)
    ap.add_argument("--out", type=Path, default=None,
                    help="Output xlsx (default: <repo>/BUYSHEET_<vendor>.xlsx)")
    ap.add_argument("--workers", type=int, default=5,
                    help="Parallel vocab-map workers (default 5)")
    args = ap.parse_args()
    if not args.pdf.exists():
        ap.error(f"PDF not found: {args.pdf}")
    if not TEMPLATE_PATH.exists():
        ap.error(f"template missing: {TEMPLATE_PATH}")
    out = args.out
    if out is None:
        pages_dir = args.pdf.with_name(args.pdf.stem + ".pages")
        vendor = None
        if pages_dir.exists():
            for jf in sorted(pages_dir.glob("*.json")):
                ctx = json.loads(jf.read_text()).get("context_after") or {}
                if ctx.get("vendor"):
                    vendor = ctx["vendor"]
                    break
        out = default_out_path(args.pdf, vendor)
    build(args.pdf, out, args.workers)


if __name__ == "__main__":
    main()
